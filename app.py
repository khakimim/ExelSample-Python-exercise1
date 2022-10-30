import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transaction.xlsx")
sheet = wb["Sheet1"]
cell = sheet.cell(1,1)

for row in range(2,sheet.max_row+1):
    price_off = sheet.cell(row,3).value * 0.9
    price_off_cell = sheet.cell(row,4)
    price_off_cell.value = price_off

values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save("transaction2.xlsx")
